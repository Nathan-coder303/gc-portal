#!/usr/bin/env python3
"""
Daily Request Tracker Upload Script
Generates a full Excel history of all user requests from Claude conversation JSONL files
and uploads it to the GC Portal Memory tab via the upload-request-tracker API.

Run manually or via cron:
  0 7 * * * /usr/bin/python3 /Users/mike/gc-portal/scripts/upload-request-tracker.py

Status overrides are persisted in:
  ~/gc-portal/scripts/request-status-overrides.json
  Format: { "<session>|<iso-timestamp>": "Fulfilled" | "Pending" }
  Edit that file manually to mark requests as fulfilled/pending across regenerations.

Requirements: pip3 install openpyxl requests
"""

import os
import json
import glob
import re
import requests
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    exit(1)

# ── Config ──────────────────────────────────────────────────────────────────
COMPANY_ID = "cmmij161r000004jm8il8bd0e"
BASE_URL = "https://gc-portal-two.vercel.app"
JSONL_DIR = os.path.expanduser("~/.claude/projects/-Users-mike")
OUTPUT_PATH = "/tmp/GC_Portal_Full_History.xlsx"
DESKTOP_PATH = os.path.expanduser("~/Desktop/GC_Portal_Full_History.xlsx")
STATUS_FILE = os.path.join(os.path.dirname(__file__), "request-status-overrides.json")

# Load CRON_SECRET from .env
env_path = os.path.expanduser("~/gc-portal/.env")
CRON_SECRET = None
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            m = re.match(r'^CRON_SECRET\s*=\s*["\']?([^"\']+)["\']?', line.strip())
            if m:
                CRON_SECRET = m.group(1)

if not CRON_SECRET:
    print("ERROR: CRON_SECRET not found in .env")
    exit(1)


# ── Status overrides ─────────────────────────────────────────────────────────
def load_status_overrides() -> dict:
    """Load manually-set status overrides from JSON sidecar file."""
    if os.path.exists(STATUS_FILE):
        try:
            with open(STATUS_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def status_key(row: dict) -> str:
    ts = row["dt"].isoformat() if row["dt"] else "unknown"
    return f"{row['session']}|{ts}"


# ── Parse JSONL files ────────────────────────────────────────────────────────
def extract_requests(jsonl_path: str) -> list[dict]:
    """Extract user messages from a JSONL conversation file."""
    requests_found = []
    try:
        with open(jsonl_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    entry = json.loads(line)
                except json.JSONDecodeError:
                    continue

                # Claude Code JSONL format: type="user", message.role="user"
                entry_type = entry.get("type")
                if entry_type != "user":
                    continue

                msg = entry.get("message", {}) or {}
                if msg.get("role") != "user":
                    continue

                # Get timestamp
                ts = entry.get("timestamp")
                dt = None
                if ts:
                    try:
                        dt = datetime.fromisoformat(ts.replace("Z", "+00:00")).replace(tzinfo=None)
                    except Exception:
                        pass

                # Get message content
                content = msg.get("content", "")
                if isinstance(content, list):
                    text_parts = []
                    for block in content:
                        if isinstance(block, dict) and block.get("type") == "text":
                            text_parts.append(block.get("text", ""))
                    content = " ".join(text_parts).strip()
                elif not isinstance(content, str):
                    content = str(content)

                content = content.strip()
                if not content or len(content) < 3:
                    continue
                # Skip very long automated context messages (system-reminder injections, etc.)
                if content.startswith("This session is being continued") or content.startswith("<system-reminder"):
                    continue
                if len(content) > 2000:
                    content = content[:2000] + "…"

                requests_found.append({
                    "dt": dt,
                    "text": content,
                    "session": os.path.basename(jsonl_path).replace(".jsonl", ""),
                })
    except Exception as e:
        print(f"  Warning: could not parse {jsonl_path}: {e}")
    return requests_found


def build_excel(all_rows: list[dict], status_overrides: dict) -> str:
    """Generate the Excel file and return the output path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Requests"

    # ── Styles ──
    gold = "C9A84C"
    dark_bg = "0D1117"
    row_bg = "1E2736"
    alt_bg = "161C26"
    fulfilled_bg = "0D2318"
    pending_bg = "2A1F00"
    header_font = Font(name="Calibri", bold=True, color=dark_bg, size=11)
    header_fill = PatternFill("solid", fgColor=gold)
    data_font = Font(name="Calibri", color="E6EDF3", size=10)
    date_font = Font(name="Calibri", color="8B949E", size=10)
    fulfilled_font = Font(name="Calibri", bold=True, color="22C55E", size=10)
    pending_font = Font(name="Calibri", bold=True, color="C9A84C", size=10)
    thin = Side(style="thin", color="30373F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(color): return PatternFill("solid", fgColor=color)

    # ── Headers ──
    headers = ["#", "Date", "Time", "Session ID", "Request / Message", "Status"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # ── Column widths ──
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["F"].width = 14

    today = datetime.now().date()

    # ── Data rows ──
    for i, row in enumerate(all_rows, 1):
        bg = row_bg if i % 2 == 1 else alt_bg
        dt = row["dt"]
        date_str = dt.strftime("%-m/%-d/%Y") if dt else "—"
        time_str = dt.strftime("%-I:%M %p") if dt else "—"

        # Determine status: check override first, then auto-detect
        key = status_key(row)
        if key in status_overrides:
            status = status_overrides[key]
        else:
            # Default: requests from today → Pending, older → Fulfilled
            if dt and dt.date() >= today:
                status = "Pending"
            else:
                status = "Fulfilled"

        values = [i, date_str, time_str, row["session"][:36], row["text"], status]
        ws.append(values)
        r = i + 1

        is_fulfilled = status == "Fulfilled"
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.border = border
            if col == 6:
                # Status column — colored background + bold colored text
                cell.fill = fill(fulfilled_bg if is_fulfilled else pending_bg)
                cell.font = fulfilled_font if is_fulfilled else pending_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.value = "✅ Fulfilled" if is_fulfilled else "⏳ Pending"
            else:
                cell.fill = fill(bg)
                cell.font = date_font if col in (2, 3, 4) else data_font
                if col == 5:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="top")

    # ── Row heights ──
    for r in range(2, len(all_rows) + 2):
        ws.row_dimensions[r].height = 42

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    today = datetime.now().strftime("%B %-d, %Y")
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Building request tracker — {today}")

    status_overrides = load_status_overrides()
    print(f"  Loaded {len(status_overrides)} status overrides from {STATUS_FILE}")

    jsonl_files = sorted(glob.glob(os.path.join(JSONL_DIR, "*.jsonl")))
    print(f"  Found {len(jsonl_files)} conversation files")

    all_rows = []
    for f in jsonl_files:
        rows = extract_requests(f)
        all_rows.extend(rows)
        if rows:
            print(f"  {os.path.basename(f)}: {len(rows)} messages")

    # Sort by date (nulls last)
    all_rows.sort(key=lambda r: r["dt"] or datetime.max)

    print(f"  Total messages: {len(all_rows)}")
    print(f"  Generating Excel → {OUTPUT_PATH}")
    build_excel(all_rows, status_overrides)

    print(f"  Uploading to {BASE_URL}/api/{COMPANY_ID}/upload-request-tracker …")
    with open(OUTPUT_PATH, "rb") as f:
        resp = requests.post(
            f"{BASE_URL}/api/{COMPANY_ID}/upload-request-tracker",
            headers={"x-cron-secret": CRON_SECRET},
            files={"file": ("GC_Portal_Full_History.xlsx", f,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=60,
        )

    if resp.ok:
        data = resp.json()
        print(f"  ✓ Uploaded — as of {data.get('date')} → {data.get('url')}")
    else:
        print(f"  ✗ Upload failed: {resp.status_code} {resp.text[:200]}")
        exit(1)

    # Also save a copy to Desktop
    import shutil
    shutil.copy2(OUTPUT_PATH, DESKTOP_PATH)
    print(f"  ✓ Saved locally → {DESKTOP_PATH}")


if __name__ == "__main__":
    main()
